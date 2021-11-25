from xls2xlsx import XLS2XLSX
import os
from os import listdir
from os.path import isfile, join
import sys
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from datetime import date


def convert (file_name):
    new_name = file_name + 'x'
    x2x = XLS2XLSX(file_name)
    x2x.to_xlsx(new_name)
    os.remove(file_name)
    return new_name

def xlsx_check (file1, file2):
    if (file1[-3:] == 'xls' or file2[-3:] == 'xls'):
        if (file1[-3:] == 'xls'):
            file1 = convert(file1)
        elif (file2[-3:] == 'xls'):
            file2 = convert(file2)
    return file1, file2

def description_builder (propType, propOperation, propLocation, propPrice, propDescription, propCharacteristics, propSQMs, propConsSQMs):
    price = int(propPrice or 0)
    currency = "${:,.0f}".format(price)
    return f"{propType.upper()} EN {propOperation.upper()} EN {propLocation.upper()} - {currency} \n{propDescription}\n\nCuenta con {propCharacteristics}\n \nTerreno: {propSQMs} m2\nConstrucción: {propConsSQMs} m2\n=================================================\n"

def compare (file1, file2):
    file1 , file2 = xlsx_check(file1,file2)
    wb1 = load_workbook(file1)
    ws1 = wb1.active
    wb2 = load_workbook(file2)
    ws2 = wb2.active
    wbF = Workbook()
    wsF = wbF.active
    f = open("Fichas técnicas.txt","w+")

    properties_old = []
    properties_new = []
    properties_added = []
    properties_deleted = []

    fill_exists = PatternFill(patternType='solid', fgColor='008000')
    fill_new = PatternFill(patternType='solid', fgColor='008000')

    for  property in ws1['A']:
        properties_old.append(property.value) 

    for  property in ws2['A']:
        properties_new.append(property.value)

    for  property in ws1['A']:
        if property.value in properties_new:
            continue
        else:
            properties_deleted.append(property.row)

    for  property in ws2['A']:
        if property.value in properties_old:
            continue
        else:
            properties_added.append(property.row)

    i = 1
    wsF['A'+str(i)] = 'Propiedades Agregadas'
    
    i = i + 1
    wsF['A'+str(i)] = ws2['E'+str(1)].value
    wsF['B'+str(i)] = ws2['AE'+str(1)].value
    wsF['C'+str(i)] = ws2['AF'+str(1)].value
    wsF['D'+str(i)] = ws2['AG'+str(1)].value
    wsF['E'+str(i)] = ws2['AI'+str(1)].value
    wsF['F'+str(i)] = ws2['AL'+str(1)].value
    wsF['G'+str(i)] = ws2['AM'+str(1)].value
    wsF['H'+str(i)] = ws2['F'+str(1)].value
    wsF['I'+str(i)] = ws2['G'+str(1)].value
    wsF['J'+str(i)] = ws2['I'+str(1)].value
    wsF['K'+str(i)] = ws2['K'+str(1)].value
    wsF['L'+str(i)] = ws2['AV'+str(1)].value
    wsF['M'+str(i)] = ws2['R'+str(1)].value
    wsF['N'+str(i)] = ws2['Z'+str(1)].value
    wsF['O'+str(i)] = ws2['X'+str(1)].value
    wsF['P'+str(i)] = ws2['Y'+str(1)].value
    wsF['Q'+str(i)] = ws2['AD'+str(1)].value
    wsF['R'+str(i)] = ws2['W'+str(1)].value
    wsF['S'+str(i)] = ws2['V'+str(1)].value
    wsF['T'+str(i)] = ws2['S'+str(1)].value
    wsF['U'+str(i)] = ws2['AQ'+str(1)].value
    wsF['V'+str(i)] = ws2['AW'+str(1)].value
    wsF['W'+str(i)] = "Tipo de Operación"
    wsF['X'+str(i)] = "Precio Operación"

    i = i + 1
    for x in properties_added:
        wsF['A'+str(i)] = ws2['E'+str(x)].value
        wsF['B'+str(i)] = ws2['AE'+str(x)].value
        wsF['C'+str(i)] = ws2['AF'+str(x)].value
        wsF['D'+str(i)] = ws2['AG'+str(x)].value
        wsF['E'+str(i)] = ws2['AI'+str(x)].value
        wsF['F'+str(i)] = ws2['AL'+str(x)].value
        wsF['G'+str(i)] = ws2['AM'+str(x)].value
        wsF['H'+str(i)] = ws2['F'+str(x)].value
        wsF['I'+str(i)] = ws2['G'+str(x)].value    
        wsF['J'+str(i)] = ws2['I'+str(x)].value
        wsF['K'+str(i)] = ws2['K'+str(x)].value
        wsF['L'+str(i)] = ws2['AV'+str(x)].value
        wsF['M'+str(i)] = ws2['R'+str(x)].value
        wsF['N'+str(i)] = ws2['Z'+str(x)].value
        wsF['O'+str(i)] = ws2['X'+str(x)].value
        wsF['P'+str(i)] = ws2['Y'+str(x)].value
        wsF['Q'+str(i)] = ws2['AD'+str(x)].value
        wsF['R'+str(i)] = ws2['W'+str(x)].value
        wsF['S'+str(i)] = ws2['V'+str(x)].value
        wsF['T'+str(i)] = ws2['S'+str(x)].value
        wsF['U'+str(i)] = ws2['AQ'+str(x)].value
        wsF['V'+str(i)] = ws2['AW'+str(x)].value
        if wsF['H'+str(i)].value == "true" and wsF['I'+str(i)].value == "false":
            wsF['W'+str(i)] = "VENTA"
            wsF['X'+str(i)] = wsF['J'+str(i)].value
        elif wsF['H'+str(i)].value == "false" and wsF['I'+str(i)].value == "true":
            wsF['W'+str(i)] = "RENTA"
            wsF['X'+str(i)] = wsF['K'+str(i)].value
        elif wsF['H'+str(i)].value == "true" and wsF['I'+str(i)].value == "true":
            wsF['W'+str(i)] = "RENTA"
            wsF['X'+str(i)] = wsF['K'+str(i)].value

        f.write(description_builder(wsF['M'+str(i)].value, wsF['W'+str(i)].value, wsF['G'+str(i)].value, wsF['X'+str(i)].value, wsF['T'+str(i)].value, wsF['U'+str(i)].value, wsF['R'+str(i)].value, wsF['S'+str(i)].value))

        i = i + 1

    i = i + 1
    
    wsF['A'+str(i)] = 'Propiedades Eliminadas'
    
    i = i + 1
    wsF['A'+str(i)] = ws2['E'+str(1)].value
    wsF['B'+str(i)] = ws2['AE'+str(1)].value
    wsF['C'+str(i)] = ws2['AF'+str(1)].value
    wsF['D'+str(i)] = ws2['AG'+str(1)].value
    wsF['E'+str(i)] = ws2['AI'+str(1)].value
    wsF['F'+str(i)] = ws2['AL'+str(1)].value
    wsF['G'+str(i)] = ws2['AM'+str(1)].value
    wsF['H'+str(i)] = ws2['F'+str(1)].value
    wsF['I'+str(i)] = ws2['G'+str(1)].value
    wsF['J'+str(i)] = ws2['I'+str(1)].value
    wsF['K'+str(i)] = ws2['K'+str(1)].value
    wsF['L'+str(i)] = ws2['AV'+str(1)].value
    wsF['M'+str(i)] = ws2['R'+str(1)].value
    wsF['N'+str(i)] = ws2['Z'+str(1)].value
    wsF['O'+str(i)] = ws2['X'+str(1)].value
    wsF['P'+str(i)] = ws2['Y'+str(1)].value
    wsF['Q'+str(i)] = ws2['AD'+str(1)].value
    wsF['R'+str(i)] = ws2['W'+str(1)].value
    wsF['S'+str(i)] = ws2['V'+str(1)].value
    wsF['T'+str(i)] = ws2['S'+str(1)].value
    wsF['U'+str(i)] = ws2['AQ'+str(1)].value
    
    i = i + 1

    for x in properties_deleted:
        wsF['A'+str(i)] = ws2['E'+str(x)].value
        wsF['B'+str(i)] = ws2['AE'+str(x)].value
        wsF['C'+str(i)] = ws2['AF'+str(x)].value
        wsF['D'+str(i)] = ws2['AG'+str(x)].value
        wsF['E'+str(i)] = ws2['AI'+str(x)].value
        wsF['F'+str(i)] = ws2['AL'+str(x)].value
        wsF['G'+str(i)] = ws2['AM'+str(x)].value
        wsF['H'+str(i)] = ws2['F'+str(x)].value
        wsF['I'+str(i)] = ws2['G'+str(x)].value
        wsF['J'+str(i)] = ws2['I'+str(x)].value
        wsF['K'+str(i)] = ws2['K'+str(x)].value
        wsF['L'+str(i)] = ws2['AV'+str(x)].value
        wsF['M'+str(i)] = ws2['R'+str(x)].value
        wsF['N'+str(i)] = ws2['Z'+str(x)].value
        wsF['O'+str(i)] = ws2['X'+str(x)].value
        wsF['P'+str(i)] = ws2['Y'+str(x)].value
        wsF['Q'+str(i)] = ws2['AD'+str(x)].value
        wsF['R'+str(i)] = ws2['W'+str(x)].value
        wsF['S'+str(i)] = ws2['V'+str(x)].value
        wsF['T'+str(i)] = ws2['S'+str(x)].value
        wsF['U'+str(i)] = ws2['AQ'+str(x)].value

        i = i + 1

    f.close()
    wbF.save('Actualizacion.xlsx')

        
def file2compare():
    mypath = os.getcwd()

    propiedades=[]
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    for  f in onlyfiles:
        files = f.split('-')
        if (files[0] == 'propiedades'):
            propiedades.append(f)

    #En caso de que no hayan archivos de propiedades?
    print('Archivos de propiedades disponibles en esta carpeta:\n')
    print(propiedades)
    return propiedades

def todayFile():
    today = date.today()
    d1 = today.strftime("%d/%m/%Y").split('/')
    fName = "propiedades-" + d1[2] + '-' + d1[1] + '-'+ d1[0] + '.xls'
    return fName

def usrReqFile(prop_files):
    print('Favor de ingresar la fecha de comparación deseada dd/mm/aaaa')
    today = input()
    d1 = today.split('/')
    fName = "propiedades-" + d1[2] + '-' + d1[1] + '-'+ d1[0] + '.xlsx'
    print(fName)
    if fName in prop_files:
        return fName
    else:
        print('Esta fecha no tiene un archivo correspondiente')

def todayFileChecker(file,prop_files):
    
    if file in prop_files:
        return file

    else:
        print('No hay archivo correspondiente para el dia de hoy')

file1 = usrReqFile(file2compare())
file2 = todayFileChecker(todayFile() , file2compare())

if (file1 != None) and (file2 != None):
    compare(file1,file2)
    print("Finished comparing")
else:
    print('NOT Comparing!')